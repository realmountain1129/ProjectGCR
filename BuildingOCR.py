import PIL
import easyocr
import os
import sys
import shutil
import errno
import re
from PIL import Image
import matplotlib.pyplot as plt
import sqlite3 as lite

CROP_PATH_1 = r"D:\\crop_buil"


################################
###  건축물대장 OCR을 위한 py  ###
################################

# easyOCR Model Load
reader = easyocr.Reader(['ko', 'en'], gpu=False)

# 0.조건 설정

upper_address = ['서울특별시','서울시','부산광역시','부산시','부산','인천광역시','인천시','인천','대구광역시','대구시','대구','광주광역시','광주시','광주','대전광역시','대전시','대전','울산광역시','울산시','울산','세종특별자치시','세종시','세종','경기도','경기','강원도','강원','충청북도','충북','충청남도','충남','전라북도','전북','전라남도','전남','경상북도','경북','경상남도','경남','제주특별자치도','제주도','제주']
error_char = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
              'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v','w', 'x', 'y', 'z',
              '1', '2', '3', '4', '5', '6', '7', '8', '9', '0', ',', '.', '/', '-', '=', '+', '!', '@', '#', '$', '%', '^', '&', '*', '(', ')']
p1 = re.compile(r'서울|부산|대구|대전|울산|광주|인천|경기|강원|충청|경상|전라*')
p2 = re.compile(r'(\d*)[-](\d{0,2})')
p3 = re.compile(r'(\d*)[-](\d{0,2})(=|-)')
p4 = re.compile(r"([가-힣]+)([0-9]+[가-힣]*)([0-9]+[가-힣]*)?") #주소 자르기
hangul = re.compile('[^ ㄱ-ㅣ가-힣]+') # 한글과 띄어쓰기를 제외한 모든 글자


# # 1-1. file list 불러오기(건축물대장)

# 1-2.엑셀에서 자료 불러오기

def call_db(filename_db):
    import openpyxl
    excel = openpyxl.load_workbook(filename_db)  # "data.xlsx"
    sheet = excel['Sheet0']  # sheet이름

    contents = {}
    register_db = []
    name_db = []
    address_db = []
    purpose_db = []
    for i in range(2, sheet.max_row + 1):
        register_number = sheet.cell(row=i, column=1).value
        register_db.append(register_number)
        name = sheet.cell(row=i, column=5).value
        name_db.append(name)
        address = sheet.cell(row=i, column=12).value
        address_db.append(address)
        # contents[name] = address #dictionary형태
        purpose = sheet.cell(row=i, column=12).value
        purpose_db.append(purpose)

    return address_db, name_db, purpose_db, register_db


# 1-3. 엑셀에서 데이터가 있는 cell 행번호 추출하기

def find_excel_num(filename, register_db):
    excel_num = 0
    for j in range(0, len(register_db)):
        if filename[:-4] == register_db[j]:
            excel_num = j
            break
    return excel_num


# 2. ocr & 자연어처리

def easyOCR(filename):  # filename = "1.jpg"

    # 이미지 불러오기
    im = PIL.Image.open(filename)

    # Doing OCR. Get bounding boxes.
    bounds = reader.readtext(filename, width_ths=1.1)
    # print(bounds) >> [텍스트좌표, 추출된텍스트, 확률]로 추출되어서 나옴

    # 1.주소 리스트
    full_address_list = []
    split_address = []
    split_address_list = []
    # 2. 이름 리스트
    name_expect = []
    name_not = []
    # 3. 용도 리스트
    purpose_list = []
    # 4. 좌표 리스트 ###
    coord_list = []

    text_num = len(bounds)
    for i in range(0, text_num):

        detect_data = bounds[i][1]  # 추출된 텍스트
        coord = bounds[i][0]

        ##주소 list##
        if i > 0 and i < 30:
            m1 = p1.match(bounds[i][1])
            m2 = p2.match(bounds[i][1])
            m3 = p3.match(bounds[i][1])
            if m1:
                full_address_list.append(bounds[i][1])
                coord_list.append(coord)
                # split처리한 주소
                split_address = detect_data.replace("ㆍ", " ").replace(".", " ").replace(":", " ").replace(",",
                                                                                                          " ").split(
                    ' ')
                for data in split_address:
                    split_address_list.append(data)  # ex. ["울산", "중구", "종가로", "323"]

            elif m2:
                full_address_list.append(bounds[i][1])
                if m3:
                    full_address_list.remove(bounds[i][1])
                else:
                    # split처리한 주소
                    split_address = detect_data.split(' ')
                    for data in split_address:
                        split_address_list.append(data)  # ex. ["울산", "중구", "종가로", "323"]

                    continue
            else:
                continue

        ##이름 list(한글로 이루어진 3글자 텍스트를 이름이라 예상하여 저장)##
        if len(detect_data) == 3:
            for char in detect_data:
                if char in error_char:
                    name_not.append(detect_data)  # 이름으로 예상되지 않는 3글자 리스트
                    break
                else:
                    name_expect.append(detect_data)  # 이름으로 예상되는 3글자 리스트 /ex. ["연면적", "주용도", "홍길동"]

        ##주택##
        if detect_data[:2] == "주택" or detect_data[:3] == "단독주" or detect_data[-2:] == "주택" or detect_data[
                                                                                              3:5] == "주택":  # 주택이 들어가는 텍스트(단독주택, 다가구주택) 추출
            purpose_list.append(detect_data)  # ex. ["단독주택", "다가구주택(5세대)", "단독주댁"] >>주택에서 한글자 오타가 굉장히 많음

    # 중복제거
    split_address_list = set(split_address_list)  # split처리한 주소에 대한 데이터 중복 제거
    name_expect = set(name_expect)  # 이름으로 예상되는 데이터 중복제거
    name_not = set(name_not)  # 이름으로 예상되지 않는 데이터 중복제거

    # 경기 = 경기도 / 대전 = 대전시 = 대전광역시 / 등 동일한 광역자치단체 이름 추가
    # ["울산", "중구", "종가로", "323"] >> ["울산", "울산광역시", "울산시", "중구", "종가로", "323"]으로 지정(추후 정확도 높이기 위해)
    full_address = []
    for data in split_address_list:

        if data == '경기' or data == '경기도':
            full_address.append('경기')
            full_address.append('경기도')
        elif data == '강원' or data == '강원도':
            full_address.append('강원')
            full_address.append('강원도')
        elif data == '충북' or data == '충청북도':
            full_address.append('충북')
            full_address.append('충청북도')
        elif data == '충남' or data == '충청남도':
            full_address.append('충남')
            full_address.append('충청남도')
        elif data == '전북' or data == '전라북도':
            full_address.append('전북')
            full_address.append('전라북도')
        elif data == '전남' or data == '전라남도':
            full_address.append('전남')
            full_address.append('전라남도')
        elif data == '경북' or data == '경상북도':
            full_address.append('경북')
            full_address.append('경상북도')
        elif data == '경남' or data == '경상남도':
            full_address.append('경남')
            full_address.append('경상남도')
        elif data == '제주특별자치도' or data == '제주도' or data == '제주':
            full_address.append('제주특별자치도')
            full_address.append('제주도')
            full_address.append('제주')
        elif data == '서울특별시' or data == '서울시' or data == '서울':
            full_address.append('서울특별시')
            full_address.append('서울시')
            full_address.append('서울')
        elif data == '세종특별자치시' or data == '세종시' or data == '세종':
            full_address.append('세종특별자치시')
            full_address.append('세종시')
            full_address.append('세종')
        elif data == '부산광역시' or data == '부산시' or data == '부산':
            full_address.append('부산광역시')
            full_address.append('부산시')
            full_address.append('부산')
        elif data == '인천광역시' or data == '인천시' or data == '인천':
            full_address.append('인천광역시')
            full_address.append('인천시')
            full_address.append('인천')
        elif data == '대구광역시' or data == '대구시' or data == '대구':
            full_address.append('대구광역시')
            full_address.append('대구시')
            full_address.append('대구')
        elif data == '광주광역시' or data == '광주시' or data == '광주':
            full_address.append('광주광역시')
            full_address.append('광주시')
            full_address.append('광주')
        elif data == '대전광역시' or data == '대전시' or data == '대전':
            full_address.append('대전광역시')
            full_address.append('대전시')
            full_address.append('대전')
        elif data == '울산광역시' or data == '울산시' or data == '울산':
            full_address.append('울산광역시')
            full_address.append('울산시')
            full_address.append('울산')
        else:
            full_address.append(data)

    # 괴정로7 >> '괴정로', '7' 로 바꾸는 과정
    full_address1 = []
    for data in full_address:
        m = p4.match(data)
        try:
            n = list(m.groups())
            try:
                n.remove(None)
                for i in n:
                    full_address1.append(i)
            except:
                for i in n:
                    full_address1.append(i)
        except:
            full_address1.append(data)
    full_address = set(full_address1)
    full_address = list(full_address)

    return name_expect, name_not, full_address_list, full_address, purpose_list, coord_list


# 3. Matching

def matching(excel_num, address_db, name_db, purpose_db,
             name_expect, name_not, full_address_list, full_address, purpose_list):  # excel_num = 해당 row 번호

    ###주소###
    address_db2 = address_db[excel_num].replace("(", "").replace(")", "").replace("ㆍ", " ").replace(".", " ").replace(
        ":", " ").replace(",", " ").split(' ')

    # 괴정로7 >> '괴정로', '7', '괴정로7'
    address_db3 = []
    for data in address_db2:
        m = p4.match(data)
        try:
            n = list(m.groups())
            try:
                n.remove(None)
                for i in n:
                    address_db3.append(i)
            except:
                for i in n:
                    address_db3.append(i)
        except:
            address_db3.append(data)
    address_db3 = set(address_db3)
    address_db3 = list(address_db3)
    if "" in address_db3:
        address_db3.remove("")

    num = len(address_db3)
    count = 0
    not_include = []  # 주소 matching할 때 일치하지 않는 데이터
    for address1 in address_db3:
        if address1 in full_address:  # full_address = 광역자치단체까지 다 넣은 주소
            count += 1
        else:
            not_include.append(address1)

    print("full_address:", full_address)  # 이미지에서 detecting한 주소 + split
    print("address_in_excel", address_db3)  # 엑셀에서 불러온 주소 / 위에 full_address에 있는 내용인지 확인할 것!
    p = int(count / num * 100)
    if p >= 70:
        result_address = "pass"
        print("address_probability:", p, "%")
        if p != 100:
            print("not_include_address:", not_include)

    else:
        result_address = "fail"
        print("address probability:", p, "%")
        print("address in image:", full_address_list)
        print("address in db:", address_db[excel_num])
        print("not include address:", not_include)

    print("." * 170)

    ###이름###
    right_name = name_db[excel_num]
    if right_name in name_expect:
        print("건축물 대장에 \"", right_name, "\" 이름이 검출됩니다")
        result_name = "pass"
    else:
        print("건축물 대장에 \"", right_name, "\" 이름이  검출되지 않습니다")
        print(">>이름이 잘못 검출되었습니까? :", name_expect)
        print(">>해당 이름이 있습니까? :", name_not)

        result_name = "fail"
    print("." * 170)

    ###용도###
    # 아래 주석된 코드는 혹시나 엑셀에 있는 내용과 정확히 일치해야만 할 때
    # right_purpose = purpose_db[excel_num]
    # if right_purpose in purpose :
    #     print("용도 : 주택")
    # else :
    #     print("용도 : 알 수 없음")

    if len(purpose_list) == 0:
        print("용도 : 알 수 없음")
        result_purpose = "fail"
    else:
        print("용도 : 주택")
        result_purpose = "pass"
    print("-" * 170)

    return result_address, result_name, result_purpose


# 4. 결과

def data_result(result_address, result_name, result_purpose, coord_list):  ###

    print("주소 :", result_address, "| 이름 :", result_name, "| 주택 :", result_purpose)

    if result_address == "pass" and result_name == "pass" and result_purpose == "pass":
        print("해당 서류의 주소, 이름, 용도가 모두 일치합니다!")
    elif result_name == "fail":
        print("해당 서류의 이름이 일치하지 않습니다!")
        return "no_name"
    elif result_address == "fail":
        print("해당 서류의 주소가 일치하지 않습니다!")
        return "no_address"
    elif result_purpose == "fail":
        print("해당 서류의 용도가 일치하지 않습니다!")
        return "no_purpose"
    else:
        print("해당 서류의 주소, 이름, 용도 중 1개 이상이 일치하지 않습니다ㅠ.")
        # 주소좌표: coord_list[0]
        # 이름좌표
        # name_coord=[[[1172,700],[2202,700],[2202,1263],[1172,1263]]]

