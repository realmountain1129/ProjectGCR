from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtCore import *
from PyQt5.QtTest import *

import sqlite3 as lite
import matplotlib.pyplot as plt

import errno
import os
import re
import shutil
from pdf2jpg import pdf2jpg
from wand.image import Image
import PIL
import easyocr
import sys

import cv2
import imutils
from pyimagesearch.transform import four_point_transform
from skimage.filters import threshold_local
import pandas as pd

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

Logo1 = resource_path("big.png")
Logo2 = resource_path("energy.png")


# UI 파일 연결
# form_class = uic.loadUiType("fifth.ui")[0]
form_class = uic.loadUiType("seventh.ui")[0]

# OCR
reader_kr = easyocr.Reader(['ko'], gpu='cuda:0')
reader_en = easyocr.Reader(['en'], gpu='cuda:0')

RESULT_PATH = r"D:\result"
DESK_PATH_1 = r"D:\desk_building"
DESK_PATH_2 = r"D:\desk_standard"
CROP_PATH_1 = r"D:\\crop_buil"
CROP_PATH_2 = r"D:\\crop_stand"

query = """CREATE TABLE stand_name
            (custnum int,
            dbData varchar[100],
            image BLOB
            )
        """

query5 = """CREATE TABLE address
            (custnum int,
            dbData varchar[100],
            image BLOB
            )
        """

query6 = """CREATE TABLE purpose
            (custnum int
            )
        """

query2 = """CREATE TABLE address
            (custnum int,
            dbData varchar[100],
            address1 BLOB,
            address2 BLOB
            )
        """

query3 = """CREATE TABLE name
            (custnum int,
            dbData varchar[100],
            nameImage BLOB
            )
        """

query4 = """CREATE TABLE capacity
            (custnum int,
            dbData varchar[100],
            capacityImage BLOB
            )
        """


class convertPdf2Image:
    def __init__(self):
        self.IMAGE_PATH = r"D:\image"
        self.CVT_PATH = r"D:\cvt"
        self.TEMP_PATH = r"D:\cvt_temp"
        self.DESK_PATH = r"D:\desk"
        self.RESULT_PATH = r"D:\result"
        self.CROP_PATH_2 = r"D:\crop_stand"
        self.CROP_PATH_1 = r"D:\crop_buil"

        # jpg 파일 저장할 폴더 & temp 저장할 폴더 생성
        # Hoxy... 그런 이름의 폴더가 있다면... 지워주기...
        try:
            if not (os.path.isdir(self.IMAGE_PATH)):
                os.makedirs(os.path.join(self.IMAGE_PATH))
            elif os.path.isdir(self.IMAGE_PATH):
                shutil.rmtree(self.IMAGE_PATH)
                os.makedirs(os.path.join(self.IMAGE_PATH))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("IMAGE 폴더를 못만들었습니다ㅠㅜ")
                raise

        try:
            if not (os.path.isdir(self.CVT_PATH)):
                os.makedirs(os.path.join(self.CVT_PATH))
            elif os.path.isdir(self.CVT_PATH):
                shutil.rmtree(self.CVT_PATH)
                os.makedirs(os.path.join(self.CVT_PATH))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("CVT 폴더를 못만들었습니다ㅠㅜ")
                raise

        try:
            if not (os.path.isdir(self.TEMP_PATH)):
                os.makedirs(os.path.join(self.TEMP_PATH))
            elif os.path.isdir(self.TEMP_PATH):
                shutil.rmtree(self.TEMP_PATH)
                os.makedirs(os.path.join(self.TEMP_PATH))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("CVT_TEMP 폴더를 못만들었습니다ㅠㅜ")

        try:
            if not (os.path.isdir(self.DESK_PATH)):
                os.makedirs(os.path.join(self.DESK_PATH))
            elif os.path.isdir(self.DESK_PATH):
                shutil.rmtree(self.DESK_PATH)
                os.makedirs(os.path.join(self.DESK_PATH))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("DESK_PATH 폴더를 못만들었습니다ㅠㅜ")

        try:
            if not (os.path.isdir(self.RESULT_PATH)):
                os.makedirs(os.path.join(self.RESULT_PATH))
            elif os.path.isdir(self.RESULT_PATH):
                shutil.rmtree(self.RESULT_PATH)
                os.makedirs(os.path.join(self.RESULT_PATH))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("RESULT_PATH 폴더를 못만들었습니다ㅠㅜ")

        try:
            if not (os.path.isdir(self.CROP_PATH_1)):
                os.makedirs(os.path.join(self.CROP_PATH_1))
            elif os.path.isdir(self.CROP_PATH_1):
                shutil.rmtree(self.CROP_PATH_1)
                os.makedirs(os.path.join(self.CROP_PATH_1))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("CROP_PATH_1 폴더를 못만들었습니다ㅠㅜ")

        try:
            if not (os.path.isdir(self.CROP_PATH_2)):
                os.makedirs(os.path.join(self.CROP_PATH_2))
            elif os.path.isdir(self.CROP_PATH_2):
                shutil.rmtree(self.CROP_PATH_2)
                os.makedirs(os.path.join(self.CROP_PATH_2))
        except OSError as e:
            if e.errno != errno.EEXIST:
                print("CROP_PATH_2 폴더를 못만들었습니다ㅠㅜ")

    # path는 시작 위치. output은 전역...

    # 파일 옮기기 (흩어진 폴더들에서 pdf(PDF)만 모아오기)
    # 1. 폴더 안 폴더의 경우
    '''
    def moveFile(self, path, file_type):
        # cnt = len(os.listdir(path))
        for dir_name in os.listdir(path):
            file_name = os.listdir(path + "\\" + dir_name)[0]
            if (file_name[-3:] == file_type.upper()) or (file_name[-3:] == file_type.lower()):
                shutil.copy(path + "\\" + dir_name + "\\" + file_name, self.IMAGE_PATH)
    '''

    # 파일 옮기기
    def moveFile(self, path, file_type):
        # cnt = len(os.listdir(path))
        for file_name in os.listdir(path):
            if (file_name[-3:] == file_type.lower()) or (file_name[-3:] == file_type.upper()):
                # print("경로 확인: ", path)
                shutil.copy(os.path.join(path, file_name), self.IMAGE_PATH)  # path + "\\" + file_name, self.IMAGE_PATH)

    # 파일 변환하기
    # 조건 : 1페이지에 필요한 서류가 있어야 함!!
    def cvtFile(self):
        for file_name in os.listdir(self.IMAGE_PATH):
            pdf2jpg.convert_pdf2jpg(self.IMAGE_PATH + "\\" + file_name, self.TEMP_PATH, dpi=200, pages="0")

    # jpg 파일 한 데 모으기
    def cltImage(self):
        for dir_name in os.listdir(self.TEMP_PATH):
            file_name = os.listdir(self.TEMP_PATH + "\\" + dir_name)[0]
            if file_name[-3:] == "jpg":
                shutil.copy(self.TEMP_PATH + "\\" + dir_name + "\\" + file_name, self.CVT_PATH)

    # 이름 바꾸기
    def chgName(self):
        for file_name in os.listdir(self.CVT_PATH):
            if file_name[-3:] == "jpg":
                os.rename(self.CVT_PATH + "\\" + file_name, self.CVT_PATH + "\\" + file_name[2:11] + ".jpg")

    # 이미지 보정하기
    def deskewImage(self):
        for file_name in os.listdir(self.CVT_PATH):
            with Image(filename=self.CVT_PATH + "\\" + file_name) as img:
                img.deskew(0.4 * img.quantum_range)
                img.type = "grayscale"
                img.save(filename=self.DESK_PATH + "\\" + file_name)

    # 작업에 사용됐던 폴더들 지우기...
    def removeDir(self):
        # 작업에 사용된 폴더들 삭제
        shutil.rmtree(self.TEMP_PATH)
        shutil.rmtree(self.IMAGE_PATH)

    def removeCvtDir(self):
        # 이미지 저장된 폴더 삭제
        shutil.rmtree(self.CVT_PATH)

    def removeDeskDirBuil(self):
        # 보정 이미지 저장된 폴더 삭제
        shutil.rmtree(self.DESK_PATH + "_building")

    def removeDeskDirStand(self):
        # 보정 이미지 저장된 폴더 삭제
        shutil.rmtree(self.DESK_PATH + "_standard")

    def removeCropDir(self):
        # 크롭 이미지 저장된 폴더 삭제
        shutil.rmtree(self.CROP_PATH_1)
        shutil.rmtree(self.CROP_PATH_2)


################################
###  건축물대장 OCR을 위한 py  ###
################################

# easyOCR Model Load
reader = easyocr.Reader(['ko', 'en'], gpu='cuda:0')

# 0.조건 설정

upper_address = ['서울특별시','서울시','부산광역시','부산시','부산','인천광역시','인천시','인천','대구광역시','대구시','대구','광주광역시','광주시','광주','대전광역시','대전시','대전','울산광역시','울산시','울산','세종특별자치시','세종시','세종','경기도','경기','강원도','강원','충청북도','충북','충청남도','충남','전라북도','전북','전라남도','전남','경상북도','경북','경상남도','경남','제주특별자치도','제주도','제주']
error_char = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
              'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v','w', 'x', 'y', 'z',
              '1', '2', '3', '4', '5', '6', '7', '8', '9', '0', ',', '.', '/', '-', '=', '+', '!', '@', '#', '$', '%', '^', '&', '*', '(', ')']
p1 = re.compile(r'서울|부산|대구|대전|울산|광주|인천|경기|강원|충청|경상|전라*')
p2 = re.compile(r'(\d*)[-](\d{0,2})')
p3 = re.compile(r'(\d*)[-](\d{0,2})(=|-)')
p4 = re.compile(r"([가-힣]+)([0-9]+[가-힣]*)([0-9]+[가-힣]*)?") #주소 자르기
hangul = re.compile('[^ ㄱ-ㅣ가-힣]+') # 한글과 띄어쓰기를 제외한 모든 글자


# # 1-1. file list 불러오기(건축물대장)

# 1-2.엑셀에서 자료 불러오기

def call_db(filename_db):
    import openpyxl
    excel = openpyxl.load_workbook(filename_db)  # "data.xlsx"
    sheet = excel['Sheet0']  # sheet이름

    contents = {}
    register_db = []
    name_db = []
    address_db = []
    purpose_db = []
    for i in range(2, sheet.max_row + 1):
        register_number = sheet.cell(row=i, column=1).value
        register_db.append(register_number)
        name = sheet.cell(row=i, column=5).value
        name_db.append(name)
        address = sheet.cell(row=i, column=12).value
        address_db.append(address)
        # contents[name] = address #dictionary형태
        purpose = sheet.cell(row=i, column=12).value
        purpose_db.append(purpose)

    return address_db, name_db, purpose_db, register_db


# 1-3. 엑셀에서 데이터가 있는 cell 행번호 추출하기

def find_excel_num(filename, register_db):
    excel_num = 0
    for j in range(0, len(register_db)):
        if filename[:-4] == register_db[j]:
            excel_num = j
            break
    return excel_num


# 2. ocr & 자연어처리

def easyOCR(filename):  # filename = "1.jpg"

    # 이미지 불러오기
    im = PIL.Image.open(filename)

    # Doing OCR. Get bounding boxes.
    bounds = reader.readtext(filename, width_ths=1.1)
    # print(bounds) >> [텍스트좌표, 추출된텍스트, 확률]로 추출되어서 나옴

    # 1.주소 리스트
    full_address_list = []
    split_address = []
    split_address_list = []
    # 2. 이름 리스트
    name_expect = []
    name_not = []
    # 3. 용도 리스트
    purpose_list = []
    # 4. 좌표 리스트 ###
    coord_list = []

    text_num = len(bounds)
    for i in range(0, text_num):

        detect_data = bounds[i][1]  # 추출된 텍스트
        coord = bounds[i][0]

        ##주소 list##
        if i > 0 and i < 30:
            m1 = p1.match(bounds[i][1])
            m2 = p2.match(bounds[i][1])
            m3 = p3.match(bounds[i][1])
            if m1:
                full_address_list.append(bounds[i][1])
                coord_list.append(coord)
                # split처리한 주소
                split_address = detect_data.replace("ㆍ", " ").replace(".", " ").replace(":", " ").replace(",",
                                                                                                          " ").split(
                    ' ')
                for data in split_address:
                    split_address_list.append(data)  # ex. ["울산", "중구", "종가로", "323"]

            elif m2:
                full_address_list.append(bounds[i][1])
                if m3:
                    full_address_list.remove(bounds[i][1])
                else:
                    # split처리한 주소
                    split_address = detect_data.split(' ')
                    for data in split_address:
                        split_address_list.append(data)  # ex. ["울산", "중구", "종가로", "323"]

                    continue
            else:
                continue

        ##이름 list(한글로 이루어진 3글자 텍스트를 이름이라 예상하여 저장)##
        if len(detect_data) == 3:
            for char in detect_data:
                if char in error_char:
                    name_not.append(detect_data)  # 이름으로 예상되지 않는 3글자 리스트
                    break
                else:
                    name_expect.append(detect_data)  # 이름으로 예상되는 3글자 리스트 /ex. ["연면적", "주용도", "홍길동"]

        ##주택##
        if detect_data[:2] == "주택" or detect_data[:3] == "단독주" or detect_data[-2:] == "주택" or detect_data[
                                                                                              3:5] == "주택":  # 주택이 들어가는 텍스트(단독주택, 다가구주택) 추출
            purpose_list.append(detect_data)  # ex. ["단독주택", "다가구주택(5세대)", "단독주댁"] >>주택에서 한글자 오타가 굉장히 많음

    # 중복제거
    split_address_list = set(split_address_list)  # split처리한 주소에 대한 데이터 중복 제거
    name_expect = set(name_expect)  # 이름으로 예상되는 데이터 중복제거
    name_not = set(name_not)  # 이름으로 예상되지 않는 데이터 중복제거

    # 경기 = 경기도 / 대전 = 대전시 = 대전광역시 / 등 동일한 광역자치단체 이름 추가
    # ["울산", "중구", "종가로", "323"] >> ["울산", "울산광역시", "울산시", "중구", "종가로", "323"]으로 지정(추후 정확도 높이기 위해)
    full_address = []
    for data in split_address_list:

        if data == '경기' or data == '경기도':
            full_address.append('경기')
            full_address.append('경기도')
        elif data == '강원' or data == '강원도':
            full_address.append('강원')
            full_address.append('강원도')
        elif data == '충북' or data == '충청북도':
            full_address.append('충북')
            full_address.append('충청북도')
        elif data == '충남' or data == '충청남도':
            full_address.append('충남')
            full_address.append('충청남도')
        elif data == '전북' or data == '전라북도':
            full_address.append('전북')
            full_address.append('전라북도')
        elif data == '전남' or data == '전라남도':
            full_address.append('전남')
            full_address.append('전라남도')
        elif data == '경북' or data == '경상북도':
            full_address.append('경북')
            full_address.append('경상북도')
        elif data == '경남' or data == '경상남도':
            full_address.append('경남')
            full_address.append('경상남도')
        elif data == '제주특별자치도' or data == '제주도' or data == '제주':
            full_address.append('제주특별자치도')
            full_address.append('제주도')
            full_address.append('제주')
        elif data == '서울특별시' or data == '서울시' or data == '서울':
            full_address.append('서울특별시')
            full_address.append('서울시')
            full_address.append('서울')
        elif data == '세종특별자치시' or data == '세종시' or data == '세종':
            full_address.append('세종특별자치시')
            full_address.append('세종시')
            full_address.append('세종')
        elif data == '부산광역시' or data == '부산시' or data == '부산':
            full_address.append('부산광역시')
            full_address.append('부산시')
            full_address.append('부산')
        elif data == '인천광역시' or data == '인천시' or data == '인천':
            full_address.append('인천광역시')
            full_address.append('인천시')
            full_address.append('인천')
        elif data == '대구광역시' or data == '대구시' or data == '대구':
            full_address.append('대구광역시')
            full_address.append('대구시')
            full_address.append('대구')
        elif data == '광주광역시' or data == '광주시' or data == '광주':
            full_address.append('광주광역시')
            full_address.append('광주시')
            full_address.append('광주')
        elif data == '대전광역시' or data == '대전시' or data == '대전':
            full_address.append('대전광역시')
            full_address.append('대전시')
            full_address.append('대전')
        elif data == '울산광역시' or data == '울산시' or data == '울산':
            full_address.append('울산광역시')
            full_address.append('울산시')
            full_address.append('울산')
        else:
            full_address.append(data)

    # 괴정로7 >> '괴정로', '7' 로 바꾸는 과정
    full_address1 = []
    for data in full_address:
        m = p4.match(data)
        try:
            n = list(m.groups())
            try:
                n.remove(None)
                for i in n:
                    full_address1.append(i)
            except:
                for i in n:
                    full_address1.append(i)
        except:
            full_address1.append(data)
    full_address = set(full_address1)
    full_address = list(full_address)

    return name_expect, name_not, full_address_list, full_address, purpose_list, coord_list


# 3. Matching

def matching(excel_num, address_db, name_db, purpose_db,
             name_expect, name_not, full_address_list, full_address, purpose_list):  # excel_num = 해당 row 번호

    ###주소###
    address_db2 = address_db[excel_num].replace("(", "").replace(")", "").replace("ㆍ", " ").replace(".", " ").replace(
        ":", " ").replace(",", " ").split(' ')

    # 괴정로7 >> '괴정로', '7', '괴정로7'
    address_db3 = []
    for data in address_db2:
        m = p4.match(data)
        try:
            n = list(m.groups())
            try:
                n.remove(None)
                for i in n:
                    address_db3.append(i)
            except:
                for i in n:
                    address_db3.append(i)
        except:
            address_db3.append(data)
    address_db3 = set(address_db3)
    address_db3 = list(address_db3)
    if "" in address_db3:
        address_db3.remove("")

    num = len(address_db3)
    count = 0
    not_include = []  # 주소 matching할 때 일치하지 않는 데이터
    for address1 in address_db3:
        if address1 in full_address:  # full_address = 광역자치단체까지 다 넣은 주소
            count += 1
        else:
            not_include.append(address1)

    print("full_address:", full_address)  # 이미지에서 detecting한 주소 + split
    print("address_in_excel", address_db3)  # 엑셀에서 불러온 주소 / 위에 full_address에 있는 내용인지 확인할 것!
    p = int(count / num * 100)
    if p >= 70:
        result_address = "pass"
        print("address_probability:", p, "%")
        if p != 100:
            print("not_include_address:", not_include)

    else:
        result_address = "fail"
        print("address probability:", p, "%")
        print("address in image:", full_address_list)
        print("address in db:", address_db[excel_num])
        print("not include address:", not_include)

    print("." * 170)

    ###이름###
    right_name = name_db[excel_num]
    if right_name in name_expect:
        print("건축물 대장에 \"", right_name, "\" 이름이 검출됩니다")
        result_name = "pass"
    else:
        print("건축물 대장에 \"", right_name, "\" 이름이  검출되지 않습니다")
        print(">>이름이 잘못 검출되었습니까? :", name_expect)
        print(">>해당 이름이 있습니까? :", name_not)

        result_name = "fail"
    print("." * 170)

    ###용도###
    # 아래 주석된 코드는 혹시나 엑셀에 있는 내용과 정확히 일치해야만 할 때
    # right_purpose = purpose_db[excel_num]
    # if right_purpose in purpose :
    #     print("용도 : 주택")
    # else :
    #     print("용도 : 알 수 없음")

    if len(purpose_list) == 0:
        print("용도 : 알 수 없음")
        result_purpose = "fail"
    else:
        print("용도 : 주택")
        result_purpose = "pass"
    print("-" * 170)

    return result_address, result_name, result_purpose


# 4. 결과

def data_result(result_address, result_name, result_purpose, coord_list):  ###

    print("주소 :", result_address, "| 이름 :", result_name, "| 주택 :", result_purpose)

    if result_address == "pass" and result_name == "pass" and result_purpose == "pass":
        print("해당 서류의 주소, 이름, 용도가 모두 일치합니다!")
    elif result_name == "fail":
        print("해당 서류의 이름이 일치하지 않습니다!")
        return "no_name"
    elif result_address == "fail":
        print("해당 서류의 주소가 일치하지 않습니다!")
        return "no_address"
    elif result_purpose == "fail":
        print("해당 서류의 용도가 일치하지 않습니다!")
        return "no_purpose"
    else:
        print("해당 서류의 주소, 이름, 용도 중 1개 이상이 일치하지 않습니다ㅠ.")
        # 주소좌표: coord_list[0]
        # 이름좌표
        # name_coord=[[[1172,700],[2202,700],[2202,1263],[1172,1263]]]


###################################
###  표준설치계약서 OCR을 위한 py  ###
###################################


def auto_crop(path):
    image = cv2.imread(path)
    ratio = 1
    orig = image.copy()
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(gray, 75, 200)

    cnts = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    cnts = sorted(cnts, key=cv2.contourArea, reverse=True)[:5]

    for c in cnts:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)

        if len(approx) == 4:
            screenCnt = approx
            break

    # Find contours of documents
    cv2.drawContours(image, [screenCnt], -1, (0, 255, 0), 2)

    # apply the four point transform to obtain a top-down
    # view of the original image
    warped = four_point_transform(orig, screenCnt.reshape(4, 2) * ratio)

    # convert the warped image to grayscale, then threshold it
    # to give it that 'black and white' paper effect
    warped = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    T = threshold_local(warped, 11, offset=10, method='gaussian')
    warped = (warped > T).astype("uint8") * 255

    warped = cv2.resize(warped, dsize=(800, 1000), interpolation=cv2.INTER_AREA)
    # show the original and scanned images
    # Apply perspective transform

    return warped


# threshold
def thresholding(image):
    return cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]


def contours_name(orig, contours, index, df, custnum, db_con):
    for cnt in contours:
        x, y, w, h = 145, 40, 100, 90  # x,y,width,height

        image_crop = orig[y:y + h, x:x + w]
        image = plt.imshow(image_crop, cmap='gray')
        plt.axis('off')
        plt.savefig(os.path.join(CROP_PATH_2, "crop.png"), bbox_inches='tight')

        image_double = cv2.pyrUp(image_crop, dstsize=(w * 2, h * 2), borderType=cv2.BORDER_DEFAULT)

        file1 = open(os.path.join(CROP_PATH_2, "crop.png"), "rb")
        img1 = file1.read()

        # Apply OCR on the cropped image
        text = reader_kr.readtext(image_double)

    try:
        if text[0][1] == df['신청자명'][index]:
            return "신청자명이 일치합니다!"
        else:
            b_crop = lite.Binary(img1)
            db_con.execute("INSERT OR REPLACE INTO name VALUES(?,?,?)", (custnum, df['신청자명'][index], b_crop))
            db_con.commit()
    except IndexError:
        db_con.execute("INSERT OR REPLACE INTO name VALUES(?,?,?)", (custnum, df['신청자명'][index], None))
        db_con.commit()


def contours_address(orig, contours, index, df, custnum, db_con):
    for cnt in contours:
        x, y, w, h = 165, 287, 320, 60  # x,y,width,height

        dst = orig.copy()
        image_crop = orig[y:y + h, x:x + w]
        image = plt.imshow(image_crop, cmap='gray')
        plt.axis('off')
        plt.savefig(os.path.join(CROP_PATH_2, "crop1.png"), bbox_inches='tight')
        file1 = open(os.path.join(CROP_PATH_2, "crop1.png"), "rb")
        img1 = file1.read()
        # cv2.imwrite('test.png', image_crop)
        b_crop = lite.Binary(img1)

        b = image_crop.tostring()

        return b_crop


def contours_address2(orig, contours, index, df, custnum, db_con):
    for cnt in contours:
        x, y, w, h = 330, 45, 470, 60  # x,y,width,height

        dst = orig.copy()
        image_crop = orig[y:y + h, x:x + w]
        image = plt.imshow(image_crop, cmap='gray')
        plt.axis('off')
        plt.savefig(os.path.join(CROP_PATH_2, "crop2.png"), bbox_inches='tight')
        file2 = open(os.path.join(CROP_PATH_2, "crop2.png"), "rb")
        img2 = file2.read()

        b_crop = lite.Binary(img2)
        b = image_crop.tostring()

        image_double = cv2.pyrUp(image_crop, dstsize=(w * 2, h * 2), borderType=cv2.BORDER_DEFAULT)

        # Apply OCR on the cropped image
        text = reader_kr.readtext(image_double)
        # return text[0][1]
        return b_crop


def contours_capacity(orig, contours, index, df, custnum, db_con):
    for cnt in contours:
        x, y, w, h = 680, 220, 90, 50  # x,y,width,height

        dst = orig.copy()
        image_crop = orig[y:y + h, x:x + w]
        image = plt.imshow(image_crop, cmap='gray')
        plt.axis('off')
        plt.savefig(os.path.join(CROP_PATH_2, "crop4.png"), bbox_inches='tight')

        # b_crop = lite.Binary(image_crop)

        image_double = cv2.pyrUp(image_crop, dstsize=(w * 2, h * 2), borderType=cv2.BORDER_DEFAULT)

        # cv2.imwrite('capacity.jpg', image_crop)

        # Apply OCR on the cropped image
        text = reader_en.readtext(image_double)
        r = re.compile("(\d+)(\s|\w)")
        m = r.match(text[0][1])

        # file4 = open(os.path.join(CROP_PATH, "crop4.png"), encoding='utf-16-le')
        # img4 = file4.read()

        try:
            if int(m.groups()[0]) == int(df['실제설치용량'][index]):
                return 'Yes'
            else:
                file4 = open(os.path.join(CROP_PATH_2, "crop4.png"), "rb")
                img4 = file4.read()

                b_crop = lite.Binary(img4)

                # b_crop = lite.Binary(img1)
                db_con.execute("INSERT INTO capacity VALUES(?,?,?)", (custnum, df['실제설치용량'][index], b_crop))
                db_con.commit()

        except AttributeError:
            return '용량이 일치하지 않습니다. 이미지와 비교해주세요! \n'
        except IndexError:
            db_con.execute("INSERT INTO capacity VALUES(?,?,?)", (custnum, df['실제설치용량'][index], None))
            db_con.commit()


# OCR 수행
def standOCR(img_path, img_list, ex_name, db_con):
    df = pd.read_excel(ex_name)
    for img in img_list:
        try:
            num = img.split('.')
            for i in range(0, len(df.index)):
                if df['참여신청번호'][i] == int(num[0]):
                    index = (df.index[i])
                else:
                    continue
            print("auto_Crop 전 : ", os.path.join(img_path, img))
            warped = auto_crop(os.path.join(img_path, img))
            threshold = cv2.threshold(warped, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            contours_name(warped, threshold, index, df, num[0], db_con)

            binary1 = contours_address(warped, threshold, index, df, num[0], db_con)
            binary2 = contours_address2(warped, threshold, index, df, num[0], db_con)

            db_con.execute("INSERT INTO address VALUES(?,?,?,?)", (num[0], df['설치주소'][index], binary1, binary2))
            db_con.commit()

            contours_capacity(warped, threshold, index, df, num[0], db_con)

        except UnboundLocalError:
            print(img + ' Error!\n')
    return


# 화면 띄우기
class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        # self.error_dialog = QMessageBox()#QErrorMessage()
        self.setupUi(self)

        # 정보 창 띄우기
        self.inform_dialog = QMessageBox()
        self.inform_dialog.setIcon(QMessageBox.Information)
        self.inform_dialog.setInformativeText("""
        프로그램 실행 조건입니다.\n
          1. 입력되는 서류(PDF)의 형태는 일정해야 합니다.
             1) 건축물대장 : 가로
             2) 표준설치계약서 : 세로\n
          2. 하나의 폴더 안에는 한 종류의 파일만 존재해야 합니다.
          """)
        self.inform_dialog.setWindowTitle("  !! 필독 !!  ")
        self.inform_dialog.setStyleSheet("""
                    QMessageBox {
                        font: 11pt "210 나무고딕 L" ;
                    }         
                """)
        self.inform_dialog.exec_()

        # 로고 띄우기
        self.label.setPixmap(self.loadImageFromFile(Logo1))
        self.label_2.setPixmap(self.loadImageFromFile(Logo2))

        # 상태 label
        self.label_3.setText("실행 전")

        # 파일 경로 초기화
        self.d_name1 = ""
        self.d_name2 = ""
        self.f_name = ""
        self.pushButton_2.clicked.connect(self.pushButtonClicked2)
        self.pushButton_3.clicked.connect(self.pushButtonClicked3)

        # FileDialog
        self.pushButton.clicked.connect(self.pushButtonClicked1)

        # Start
        self.pushButton_4.clicked.connect(self.startButtonClicked)

    # 이미지 로드
    def loadImageFromFile(self, path):
        self.qPixmapFileVar = QPixmap()
        self.qPixmapFileVar.load(path)
        return self.qPixmapFileVar

    # 각 서류 경로 불러오기
    def pushButtonClicked2(self):
        self.d_name1 = QFileDialog.getExistingDirectory(self)
        print("건축물대장 경로 : ", self.d_name1)
        if len(self.d_name1) >= 3:
            self.label_3.setText("건축물대장 입력 완료")

    def pushButtonClicked3(self):
        self.d_name2 = QFileDialog.getExistingDirectory(self)
        print("표준설치계약서 경로 : ", self.d_name2)
        if len(self.d_name2) >= 3:
            self.label_3.setText("표준설치계약서 입력 완료")

    def pushButtonClicked1(self):
        self.f_name = QFileDialog.getOpenFileName(self, filter="Excel (*.xlsx *.xls)")
        # 불러온 파일 이름 표시...
        # self.label.setText(f_name[0])
        print(self.f_name[0], "불러옴")
        if len(self.f_name[0]) >= 3:
            self.label_3.setText("Excel 입력 완료")

    def startButtonClicked(self):
        # if (self.d_name1 == "") or (self.d_name2 == "") or (self.f_name == ""):
        if (len(self.d_name1) < 3) or (len(self.d_name2) < 3) or (len(self.f_name[0]) < 3):
            error_dialog = QMessageBox()
            error_dialog.setIcon(QMessageBox.Critical)
            error_dialog.setInformativeText("파일 입력을 확인하세요!")  # showMessage("파일이 입력되지 않았습니다!")
            error_dialog.setWindowTitle("!! 파일이 부족해 !!")
            error_dialog.setStyleSheet("""
                                QMessageBox {
                                    font: 11pt "210 나무고딕 L" ;
                                }         
                            """)
            error_dialog.exec_()
        else:
            # Pdf2Jpg 작업 시작
            self.label_3.setText("PDF => JPG 변환")
            QTest.qWait(1000)

            cvt_1 = convertPdf2Image()
            cvt_1.moveFile(self.d_name1, "pdf")
            cvt_1.cvtFile()
            cvt_1.cltImage()
            cvt_1.chgName()
            cvt_1.deskewImage()
            try:
                os.rename(cvt_1.DESK_PATH, DESK_PATH_1)
            except FileExistsError:
                shutil.rmtree(DESK_PATH_1)
                os.rename(cvt_1.DESK_PATH, DESK_PATH_1)
            print("건축물대장 변환 완료!")

            self.label_3.setText("건축물대장 변환 완료!")
            QTest.qWait(1000)

            cvt_1.removeDir()
            cvt_1.removeCvtDir()

            cvt_2 = convertPdf2Image()
            cvt_2.moveFile(self.d_name2, "pdf")
            cvt_2.cvtFile()
            cvt_2.cltImage()
            cvt_2.chgName()
            cvt_2.deskewImage()
            try:
                os.rename(cvt_2.DESK_PATH, DESK_PATH_2)
            except FileExistsError:
                shutil.rmtree(DESK_PATH_2)
                os.rename(cvt_2.DESK_PATH, DESK_PATH_2)

            QApplication.processEvents()

            print("표준설치계약서 변환 완료!")
            self.label_3.setText("표준설치계약서 변환 완료!")
            QTest.qWait(1000)

            cvt_2.removeDir()
            cvt_2.removeCvtDir()

            print("OCR 시작!")
            self.label_3.setText("건축물대장 OCR")
            QTest.qWait(1000)

            # 결과 저장 DB
            buil = lite.connect(os.path.join(RESULT_PATH, "건축물대장_unpass.db"))
            stand = lite.connect(os.path.join(RESULT_PATH, "표준설치계약서_unpass.db"))

            # DB 스키마 생성
            buil.execute(query)
            buil.execute(query5)
            buil.execute(query6)

            stand.execute(query2)
            stand.execute(query3)
            stand.execute(query4)

            buil.commit()
            stand.commit()

            # Excel 파일 로드
            address_db, name_db, purpose_db, register_db = call_db(self.f_name[0])

            name_coord = [[[1172, 700], [2202, 700], [2202, 1263], [1172, 1263]]]

            # 건축물대장 OCR 작업 시작
            filename_list = os.listdir(DESK_PATH_1)

            for filename in filename_list:
                excel_num = find_excel_num(filename, register_db)  # 1-3. 엑셀 해당 cell의 행 출력 ( 2 >> 실제로 4번째행, 3 >> 5 ,,,)
                print("<", filename, ">")

                # 2. ocr & 자연어처리
                name_expect, name_not, full_address_list, full_address, purpose_list, coord_list = easyOCR(
                    os.path.join(DESK_PATH_1, filename))
                # 이름으로예상되는/예상되지않는/full주소/split처리한주소/용도

                # 3. Matching
                result_address, result_name, result_purpose = matching(excel_num,
                                                                       address_db, name_db, purpose_db,
                                                                       name_expect, name_not, full_address_list,
                                                                       full_address, purpose_list)

                # 이미지 로드
                buil_im = PIL.Image.open(os.path.join(DESK_PATH_1, filename))

                # 4. 결과
                if result_address == "pass" and result_name == "pass" and result_purpose == "pass":
                    print("해당 서류의 주소, 이름, 용도가 모두 일치합니다!")

                # 이름 INSERT
                elif result_name == "fail":
                    print("해당 서류의 이름이 일치하지 않습니다!")
                    img = buil_im.crop(
                        (name_coord[0][0][0], name_coord[0][0][1], name_coord[0][2][0], name_coord[0][2][1]))
                    image = plt.imshow(img, cmap='gray')
                    plt.axis('off')
                    plt.savefig(os.path.join(CROP_PATH_2, "crop_name.png"), bbox_inches='tight')

                    file1 = open(os.path.join(CROP_PATH_2, "crop_name.png"), "rb")
                    img1 = file1.read()
                    binary1 = lite.Binary(img1)

                    buil.execute("INSERT INTO stand_name VALUES(?,?,?)", [filename[0:9], name_db[excel_num], binary1])
                    buil.commit()

                # 주소 INSERT
                elif result_address == "fail":
                    print("해당 서류의 주소가 일치하지 않습니다!")
                    img2 = buil_im.crop(
                        (coord_list[0][0][0], coord_list[0][0][1], coord_list[0][2][0], coord_list[0][2][1]))
                    image2 = plt.imshow(img2, cmap='gray')
                    plt.axis('off')
                    plt.savefig(os.path.join(CROP_PATH_2, "crop_address.png"), bbox_inches='tight')

                    file2 = open(os.path.join(CROP_PATH_2, "crop_address.png"), "rb")
                    img2 = file2.read()
                    binary2 = lite.Binary(img2)

                    buil.execute("INSERT INTO address VALUES(?,?,?)", [filename[0:9], address_db[excel_num], binary2])
                    buil.commit()

                # 용도 INSERT ?
                elif result_purpose == "fail":
                    print("해당 서류의 용도가 일치하지 않습니다!")
                    buil.execute("INSERT INTO purpose VALUES(?)", [filename[0:9]])
                    buil.commit()
                else:
                    print("다 틀렸어!")

                data_result(result_address, result_name, result_purpose, coord_list)

            self.label_3.setText("건축물대장 OCR 끌!")
            QTest.qWait(2000)

            # 표준설치계약서 OCR 작업 시작
            self.label_3.setText("표준설치계약서 OCR")
            QTest.qWait(1000)

            filename_list2 = os.listdir(DESK_PATH_2)
            standOCR(DESK_PATH_2, filename_list2, self.f_name[0], stand)

            self.label_3.setText("표준설치계약서 OCR 끝!")
            QTest.qWait(1000)

            buil.close()
            stand.close()

            # 생성된 폴더 삭제하는 코드
            # 인데 PIL open 때문에 WinError 32 뜸
            # try 써서 처리할 사람만 살리길..
            # cvt_1.removeDeskDirBuil()
            # cvt_2.removeDeskDirStand()
            # cvt_1.removeCropDir()

            self.label_3.setText("DB에 UNPASS 결과가\n적재되었습니다.")
            QTest.qWait(1000)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
